import os
import re
import openpyxl
from sqlalchemy.orm import Session
from app.core.database import SessionLocal, engine, Base
from app.core.security import get_password_hash
from app.models.user import User, RoleEnum
from app.models.role import Role
from app.models.announcement import Announcement
from app.models.event import Event
from app.models.document import Document
from app.models.ticket import Ticket, TicketMessage
from app.models.schulung import TrainingDocument
from app.services.role_service import seed_default_roles

# Position mappings for each user
POSITIONS_MAP = {
    'Anja Knoll': 'Geschäftsführerin',
    'Anas Guist': 'Leitung Geschäftsentwicklung',
    'Cagla Karayigit': 'Referentin Geschäftsentwicklung',
    'Susanne Merten': 'Empfang & Zentrale Rezeption',
    'Andreas Walker': 'Vertriebsleitung',
    'Petra Petersen': 'Vertriebsassistenz',
    'Andreas Liebow': 'Technischer Vertrieb',
    'Stefan Meyer': 'Projektvertrieb & Kundenbetreuung',
    'Kamel Al Daher': 'Vertriebsberater',
    'Beatrix Kopczak': 'Leitung Qualitätskontrolle & QS',
    'Oja Morina Cal': 'Technische Leitung & CAD',
    'Jan Fischer': 'Bauingenieur & Statik',
    'Andreas Braun': 'CAD-Konstrukteur',
    'Ingrid Müller': 'Technische Zeichnerin',
    'Christiane Benz': 'CAD-Konstrukteurin',
    'Frank Beutling': 'Technischer Konstrukteur',
    'Ryan Würfel': 'Konstruktionsingenieur',
    'Ane Steinmetz (Azubi)': 'Auszubildende Bauzeichnerin',
    'Ahmad Quddosy': 'Konstrukteur & Statik',
    'Dani Daher': 'Technischer Zeichner',
    'Cihad Sözen': 'Konstrukteur',
    'Diana Moskalyk (Azubi)': 'Auszubildende Bauzeichnerin',
    'Barbara Peters': 'Leitung Finanzbuchhaltung',
    'Steffen Martsch': 'Finanzbuchhalter',
    'Rodica Petrean': 'Finanzbuchhalterin',
    'Moritz Thorn': 'Debitorenbuchhaltung',
    'Matthias Grade': 'Produktionsleiter & Planung',
    'Carmen Pietsch': 'Produktionsplanung & Steuerung',
    'Mario Köcher': 'Fertigungsdisponent',
    'Ingo Thiele': 'Leitung Auftragsabwicklung',
    'Franko Pade': 'Baustellen-Disposition',
    'Haci Cal': 'Transport- & Logistikabwicklung',
    'Martin Scheffler': 'Auftragsabwickler',
    'Jenny Rudolph': 'Sachbearbeiterin Abwicklung',
    'Torsten Anton': 'Baustellen-Logistik & Disposition',
    'Robert Kuhaupt': 'Senior IT-Administrator & Security',
    'Humbert Senf': 'IT-Leiter & SuperAdmin',
}

# Role slug mappings based on Excel department
DEPT_ROLE_MAP = {
    'Geschäftsführung': 'MANAGEMENT',
    'Geschäftsentwicklung': 'BUSINESS_DEV',
    'Rezeption': 'RECEPTION',
    'Vertriebsabteilung': 'SALES',
    'Kontrolle': 'CONTROLLING_QS',
    'Technik': 'TECHNIK',
    'Buchhaltung': 'ACCOUNTING',
    'Produktion \\ Planung': 'PRODUKTION',
    'Abwicklung': 'ABWICKLUNG',
    'IT \\ SuperAdmin': 'IT_ADMIN',
}

def migrate_users():
    db: Session = SessionLocal()
    try:
        print("1. Seeding / Updating department roles in DB...")
        seed_default_roles(db)

        roles = db.query(Role).all()
        roles_by_slug = {r.slug: r.id for r in roles}
        print(f"Available roles: {list(roles_by_slug.keys())}")

        print("\n2. Preserving SuperAdmin Humbert Senf and updating references...")
        humbert = db.query(User).filter(User.email == 'h.senf@tinglev.de').first()
        if not humbert:
            print("Creating Humbert Senf as SuperAdmin...")
            humbert = User(
                email="h.senf@tinglev.de",
                first_name="Humbert",
                last_name="Senf",
                full_name="Humbert Senf",
                hashed_password=get_password_hash("Passwort123!"),
                role="ADMIN",
                custom_role_id=roles_by_slug.get("ADMIN"),
                department="IT \\ SuperAdmin",
                position="IT-Leiter & SuperAdmin (HUSE)",
                phone="+49 33439 86-245",
                mobile="0162 / 25 66 144",
                location="Werk Tinglev",
                is_active=True
            )
            db.add(humbert)
            db.flush()
        else:
            humbert.role = "ADMIN"
            humbert.custom_role_id = roles_by_slug.get("ADMIN")
            humbert.is_active = True

        # Reassign all existing foreign key records to Humbert (id=humbert.id)
        db.query(Announcement).filter(Announcement.author_id != humbert.id).update({Announcement.author_id: humbert.id})
        db.query(Event).filter(Event.created_by_id != humbert.id).update({Event.created_by_id: humbert.id})
        db.query(Document).filter(Document.uploaded_by_id != humbert.id).update({Document.uploaded_by_id: humbert.id})
        db.query(Ticket).filter(Ticket.ersteller_id != humbert.id).update({Ticket.ersteller_id: humbert.id})
        db.query(Ticket).filter(Ticket.zugewiesen_an_id != humbert.id, Ticket.zugewiesen_an_id != None).update({Ticket.zugewiesen_an_id: humbert.id})
        db.query(TicketMessage).filter(TicketMessage.autor_id != humbert.id).update({TicketMessage.autor_id: humbert.id})
        db.query(TrainingDocument).filter(TrainingDocument.uploaded_by != humbert.id).update({TrainingDocument.uploaded_by: humbert.id})
        db.commit()

        print("\n3. Deleting all previous demo users (except Humbert Senf)...")
        deleted_count = db.query(User).filter(User.id != humbert.id).delete(synchronize_session=False)
        db.commit()
        print(f"Deleted {deleted_count} old demo users.")

        print("\n4. Reading Users.xlsx...")
        wb = openpyxl.load_workbook(r'c:\Users\Humbert\Desktop\intranet-corp\Users.xlsx')
        ws = wb.active

        excel_users = []
        current_dept = None

        for r in range(1, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            col0 = str(vals[0]).strip() if vals[0] is not None else ''
            col1 = str(vals[1]).strip() if vals[1] is not None else ''

            if col0.lower() == 'rolle':
                current_dept = col1
                continue

            if col0.lower() == 'name' and col1:
                clean_name = col1.replace('\xa0', ' ').strip()
                int_phone = str(vals[2]).replace('\xa0', ' ').strip() if vals[2] is not None else ''
                if int_phone == 'None': int_phone = ''
                
                email = str(vals[3]).replace('\xa0', ' ').strip().lower() if vals[3] is not None else ''
                if email == 'None': email = ''
                
                mob = str(vals[4]).replace('\xa0', ' ').strip() if vals[4] is not None else ''
                if mob == 'None': mob = ''
                
                abbr = str(vals[5]).replace('\xa0', ' ').strip().upper() if vals[5] is not None else ''
                if abbr == 'None': abbr = ''

                excel_users.append({
                    'name': clean_name,
                    'department': current_dept,
                    'internal_phone': int_phone,
                    'email': email,
                    'mobile': mob,
                    'abbreviation': abbr
                })

        print(f"Found {len(excel_users)} users in Users.xlsx.")

        # Default password for all corporate users
        default_pwd_hash = get_password_hash("Passwort123!")

        created_users_map = {
            'Humbert Senf': humbert
        }

        print("\n5. Creating user accounts in DB...")
        for u in excel_users:
            full_name = u['name']
            if full_name == 'Humbert Senf':
                # Already preserved as SuperAdmin anchor
                continue

            # Remove (Azubi) from name split
            name_for_split = full_name.replace('(Azubi)', '').strip()
            name_parts = name_for_split.split()
            first_name = name_parts[0] if len(name_parts) > 0 else full_name
            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

            dept = u['department']
            role_slug = DEPT_ROLE_MAP.get(dept, 'EMPLOYEE')
            role_id = roles_by_slug.get(role_slug, roles_by_slug.get('EMPLOYEE'))

            # External Phone Number (+49 33439 86- Durchwahl)
            int_p = u['internal_phone']
            if int_p:
                first_dw = int_p.split('/')[0].strip()
                external_phone = f"+49 33439 86-{first_dw}"
            else:
                external_phone = "+49 33439 86-0"

            base_pos = POSITIONS_MAP.get(full_name, 'Mitarbeiter')
            abbr = u['abbreviation']
            pos_with_abbr = f"{base_pos} ({abbr})" if abbr else base_pos

AVATARS_MAP = {
    'Anja Knoll': 'https://images.unsplash.com/photo-1573496359142-b8d87734a5a2?w=250&auto=format&fit=crop&q=80',
    'Anas Guist': 'https://images.unsplash.com/photo-1560250097-0b93528c311a?w=250&auto=format&fit=crop&q=80',
    'Cagla Karayigit': 'https://images.unsplash.com/photo-1580489944761-15a19d654956?w=250&auto=format&fit=crop&q=80',
    'Susanne Merten': 'https://images.unsplash.com/photo-1544005313-94ddf0286df2?w=250&auto=format&fit=crop&q=80',
    'Andreas Walker': 'https://images.unsplash.com/photo-1519085360753-af0119f7cbe7?w=250&auto=format&fit=crop&q=80',
    'Petra Petersen': 'https://images.unsplash.com/photo-1573497019940-1c28c88b4f3e?w=250&auto=format&fit=crop&q=80',
    'Andreas Liebow': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=250&auto=format&fit=crop&q=80',
    'Stefan Meyer': 'https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=250&auto=format&fit=crop&q=80',
    'Kamel Al Daher': 'https://images.unsplash.com/photo-1472099645785-5658abf4ff4e?w=250&auto=format&fit=crop&q=80',
    'Beatrix Kopczak': 'https://images.unsplash.com/photo-1567532939604-b6b5b0db2604?w=250&auto=format&fit=crop&q=80',
    'Oja Morina Cal': 'https://images.unsplash.com/photo-1573497019418-b400bb3ab074?w=250&auto=format&fit=crop&q=80',
    'Jan Fischer': 'https://images.unsplash.com/photo-1522075469751-3a6694fb2f61?w=250&auto=format&fit=crop&q=80',
    'Andreas Braun': 'https://images.unsplash.com/photo-1506794778202-cad84cf45f1d?w=250&auto=format&fit=crop&q=80',
    'Ingrid Müller': 'https://images.unsplash.com/photo-1551836022-d5d88e9218df?w=250&auto=format&fit=crop&q=80',
    'Christiane Benz': 'https://images.unsplash.com/photo-1598550874175-4d0ef436c909?w=250&auto=format&fit=crop&q=80',
    'Frank Beutling': 'https://images.unsplash.com/photo-1492562080023-ab3db95bfbce?w=250&auto=format&fit=crop&q=80',
    'Ryan Würfel': 'https://images.unsplash.com/photo-1539571696357-5a69c17a67c6?w=250&auto=format&fit=crop&q=80',
    'Ane Steinmetz (Azubi)': 'https://images.unsplash.com/photo-1517841905240-472988babdf9?w=250&auto=format&fit=crop&q=80',
    'Ahmad Quddosy': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=250&auto=format&fit=crop&q=80',
    'Dani Daher': 'https://images.unsplash.com/photo-1519085360753-af0119f7cbe7?w=250&auto=format&fit=crop&q=80',
    'Cihad Sözen': 'https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=250&auto=format&fit=crop&q=80',
    'Diana Moskalyk (Azubi)': 'https://images.unsplash.com/photo-1534528741775-53994a69daeb?w=250&auto=format&fit=crop&q=80',
    'Barbara Peters': 'https://images.unsplash.com/photo-1573496359142-b8d87734a5a2?w=250&auto=format&fit=crop&q=80',
    'Steffen Martsch': 'https://images.unsplash.com/photo-1522075469751-3a6694fb2f61?w=250&auto=format&fit=crop&q=80',
    'Rodica Petrean': 'https://images.unsplash.com/photo-1567532939604-b6b5b0db2604?w=250&auto=format&fit=crop&q=80',
    'Moritz Thorn': 'https://images.unsplash.com/photo-1472099645785-5658abf4ff4e?w=250&auto=format&fit=crop&q=80',
    'Matthias Grade': 'https://images.unsplash.com/photo-1560250097-0b93528c311a?w=250&auto=format&fit=crop&q=80',
    'Carmen Pietsch': 'https://images.unsplash.com/photo-1580489944761-15a19d654956?w=250&auto=format&fit=crop&q=80',
    'Mario Köcher': 'https://images.unsplash.com/photo-1506794778202-cad84cf45f1d?w=250&auto=format&fit=crop&q=80',
    'Ingo Thiele': 'https://images.unsplash.com/photo-1519085360753-af0119f7cbe7?w=250&auto=format&fit=crop&q=80',
    'Franko Pade': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?w=250&auto=format&fit=crop&q=80',
    'Haci Cal': 'https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=250&auto=format&fit=crop&q=80',
    'Martin Scheffler': 'https://images.unsplash.com/photo-1492562080023-ab3db95bfbce?w=250&auto=format&fit=crop&q=80',
    'Jenny Rudolph': 'https://images.unsplash.com/photo-1544005313-94ddf0286df2?w=250&auto=format&fit=crop&q=80',
    'Torsten Anton': 'https://images.unsplash.com/photo-1522075469751-3a6694fb2f61?w=250&auto=format&fit=crop&q=80',
    'Robert Kuhaupt': 'https://images.unsplash.com/photo-1539571696357-5a69c17a67c6?w=250&auto=format&fit=crop&q=80',
    'Humbert Senf': 'https://images.unsplash.com/photo-1534528741775-53994a69daeb?w=250&auto=format&fit=crop&q=80',
}

            user_obj = User(
                email=u['email'],
                first_name=first_name,
                last_name=last_name,
                full_name=full_name,
                hashed_password=default_pwd_hash,
                role=role_slug,
                custom_role_id=role_id,
                department=dept,
                position=pos_with_abbr,
                avatar_url=AVATARS_MAP.get(full_name),
                phone=external_phone,
                mobile=u['mobile'] if u['mobile'] else None,
                location="Werk Tinglev",
                is_active=True
            )
            db.add(user_obj)
            db.flush()
            created_users_map[full_name] = user_obj

        # 6. Establish Organizational Hierarchy (Supervisors)
        # CEO / Managing Director: Anja Knoll (Root)
        ceo = created_users_map.get('Anja Knoll')
        if ceo:
            ceo.supervisor_id = None

            # Department heads report to CEO Anja Knoll
            dept_heads = [
                'Anas Guist',        # Geschäftsentwicklung
                'Susanne Merten',    # Rezeption
                'Andreas Walker',    # Vertrieb
                'Beatrix Kopczak',   # Kontrolle
                'Oja Morina Cal',    # Technik
                'Barbara Peters',    # Buchhaltung
                'Matthias Grade',    # Produktion
                'Ingo Thiele',       # Abwicklung
                'Robert Kuhaupt',    # IT
                'Humbert Senf',      # IT
            ]

            for head_name in dept_heads:
                head_user = created_users_map.get(head_name)
                if head_user:
                    head_user.supervisor_id = ceo.id

            # Department staff report to their respective department head
            dept_assignments = {
                'Anas Guist': ['Cagla Karayigit'],
                'Andreas Walker': ['Petra Petersen', 'Andreas Liebow', 'Stefan Meyer', 'Kamel Al Daher'],
                'Oja Morina Cal': [
                    'Jan Fischer', 'Andreas Braun', 'Ingrid Müller', 'Christiane Benz',
                    'Frank Beutling', 'Ryan Würfel', 'Ane Steinmetz (Azubi)', 'Ahmad Quddosy',
                    'Dani Daher', 'Cihad Sözen', 'Diana Moskalyk (Azubi)'
                ],
                'Barbara Peters': ['Steffen Martsch', 'Rodica Petrean', 'Moritz Thorn'],
                'Matthias Grade': ['Carmen Pietsch', 'Mario Köcher'],
                'Ingo Thiele': ['Franko Pade', 'Haci Cal', 'Martin Scheffler', 'Jenny Rudolph', 'Torsten Anton'],
            }

            for manager_name, staff_list in dept_assignments.items():
                manager = created_users_map.get(manager_name)
                if manager:
                    for staff_name in staff_list:
                        staff = created_users_map.get(staff_name)
                        if staff:
                            staff.supervisor_id = manager.id

        db.commit()

        print("\n=== SUMMARY OF USERS IN DATABASE ===")
        all_users = db.query(User).all()
        for idx, usr in enumerate(all_users, 1):
            sup_name = usr.supervisor.full_name if usr.supervisor else 'Keine (Top-Level)'
            print(f"{idx:2d}. [ID {usr.id:2d}] {usr.full_name:<28} | Role: {usr.role:<14} | Dept: {usr.department:<22} | Tel: {usr.phone:<18} | Vorgesetzter: {sup_name}")

        print(f"\nTotal users in DB: {len(all_users)} (Carlos Mendoza + 37 corporate employees)")

    except Exception as e:
        db.rollback()
        print(f"Error during migration: {e}")
        raise e
    finally:
        db.close()

if __name__ == '__main__':
    migrate_users()
