import openpyxl, json

wb = openpyxl.load_workbook(r'c:\Users\Humbert\Desktop\intranet-corp\Users.xlsx')
ws = wb.active

current_department = None
users = []

for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    col0 = str(vals[0]).strip() if vals[0] is not None else ''
    col1 = str(vals[1]).strip() if vals[1] is not None else ''
    
    if col0.lower() == 'rolle':
        current_department = col1
        continue
    
    if col0.lower() == 'name' and col1:
        name = col1.replace('\xa0', ' ').strip()
        internal_phone = str(vals[2]).replace('\xa0', ' ').strip() if vals[2] is not None else ''
        if internal_phone == 'None': internal_phone = ''
        email = str(vals[3]).replace('\xa0', ' ').strip() if vals[3] is not None else ''
        if email == 'None': email = ''
        mobile = str(vals[4]).replace('\xa0', ' ').strip() if vals[4] is not None else ''
        if mobile == 'None': mobile = ''
        abbreviation = str(vals[5]).replace('\xa0', ' ').strip() if vals[5] is not None else ''
        if abbreviation == 'None': abbreviation = ''
        
        users.append({
            'name': name,
            'department': current_department,
            'internal_phone': internal_phone,
            'email': email,
            'mobile': mobile,
            'abbreviation': abbreviation
        })

print(f'Total parsed users: {len(users)}')
dept_counts = {}
for u in users:
    d = u['department']
    dept_counts[d] = dept_counts.get(d, 0) + 1

print('\nDepartments found:')
for d, count in dept_counts.items():
    print(f'  - {d}: {count} users')

print('\nDetailed users:')
for idx, u in enumerate(users, 1):
    print(f"{idx:2d}. [{u['department']}] {u['name']} <{u['email']}> | Durchwahl: {u['internal_phone']} | Mobil: {u['mobile']} | Kürzel: {u['abbreviation']}")
